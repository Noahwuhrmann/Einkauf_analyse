#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Dec  4 2024

@author: Noah Wuhrmann
"""

# V23

import pandas as pd
import streamlit as st
import re
import traceback
from io import BytesIO
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter


def normalize_column_name(col):
    if pd.isna(col):
        return ""
    if isinstance(col, float) and col.is_integer():
        return str(int(col))
    return str(col).strip()


# Set the theme to a dark mode
st.markdown(
    """
    <style>
    body {
        background-color: black;
        color: white;
    }
    .stApp {
        background-color: black;
    }
    h1 {
        color: white;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# App Title
st.title("Einkaufsanalyse – Taurus Sports AG")

# Add the logo with some spacing
st.image("absolut_bild.jpg", use_container_width=True)
st.markdown("<br>", unsafe_allow_html=True)

# File upload section
uploaded_file = st.file_uploader("Bitte eine Excel-Datei hochladen", type=["xlsx"])

if uploaded_file:
    st.success("Die Datei wurde erfolgreich hochgeladen.")
    try:
        # Read Excel file into DataFrame
        df = pd.read_excel(uploaded_file, engine="openpyxl")
        df.columns = [normalize_column_name(col) for col in df.columns]

        # Define fixed columns
        fixed_columns = ["Bestand", "Artikelname", "Artikelgruppe", "KatalogNr", "Artikel", "Kollektion", "Netto"]

        # Identify month columns
        month_pattern = re.compile(r"^(Jan|Feb|Mär|Mrz|Apr|Mai|Jun|Jul|Aug|Sep|Okt|Nov|Dez) [0-9]{4}$")
        month_columns = [col for col in df.columns if month_pattern.match(col)]

        # Identify year columns
        year_pattern = re.compile(r"^[0-9]{4}$")
        year_columns = [col for col in df.columns if year_pattern.match(col)]

        # Combine all relevant columns
        target_columns = fixed_columns + year_columns + month_columns

        # Filter DataFrame to retain only relevant columns
        filtered_df = df[[col for col in target_columns if col in df.columns]].copy()

        # Move "Netto" to the last position
        if "Netto" in filtered_df.columns:
            netto_col = filtered_df.pop("Netto")
            filtered_df["Netto"] = netto_col

        # Reorder "Kollektion" to be after "Artikel"
        if "Kollektion" in filtered_df.columns and "Artikel" in filtered_df.columns:
            kollektion_col = filtered_df.pop("Kollektion")
            artikel_index = filtered_df.columns.get_loc("Artikel") + 1
            left = filtered_df.iloc[:, :artikel_index]
            right = filtered_df.iloc[:, artikel_index:]
            filtered_df = pd.concat([left, kollektion_col, right], axis=1)

        # Drop rows where all relevant columns are zero
        relevant_columns = [col for col in ["Bestand"] + month_columns if col in filtered_df.columns]
        if relevant_columns:
            filtered_df = filtered_df[~((filtered_df[relevant_columns] == 0).all(axis=1))]

        # User input for output filename
        output_filename = st.text_input(
            "Bitte geben Sie einen Namen für die exportierte Datei ein (ohne .xlsx):",
            value="Exportierte_Datei"
        )

        if st.button("Excel-Datei exportieren"):
            # Create an Excel file in memory with formatting
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                filtered_df.to_excel(writer, index=False, sheet_name="Export")
                worksheet = writer.sheets["Export"]

                # Apply styles
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
                alignment_center = Alignment(horizontal="center", vertical="center")
                alignment_left = Alignment(horizontal="left", vertical="center")
                alignment_right = Alignment(horizontal="right", vertical="center")
                calibri_11 = Font(name="Calibri", size=11)

                # Format header row
                for col_num, col in enumerate(filtered_df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment_center

                # Format columns
                for col_num, col_name in enumerate(filtered_df.columns, 1):
                    col_letter = get_column_letter(col_num)
                    column_as_text = filtered_df.iloc[:, col_num - 1].map(
                        lambda x: "" if pd.isna(x) else str(x)
                    )
                    max_cell_len = column_as_text.map(len).max()
                    max_header_len = len(str(col_name))
                    worksheet.column_dimensions[col_letter].width = max(
                        15, max(max_cell_len, max_header_len) + 2
                    )

                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=col_num)
                        cell.font = calibri_11

                        if col_name == "Netto":
                            cell.number_format = "#,##0.00"
                            if isinstance(cell.value, (int, float)) and cell.value < 0:
                                cell.font = Font(color="FF0000")

                        elif col_name in ["Bestand", "KatalogNr"]:
                            cell.alignment = alignment_left

                        elif col_name not in ["Artikelname", "Artikelgruppe", "KatalogNr", "Artikel", "Kollektion"]:
                            cell.alignment = alignment_right

                # Enable autofilter
                worksheet.auto_filter.ref = worksheet.dimensions

            # Save the file to output and allow download
            output.seek(0)
            st.success(f"Die Datei '{output_filename}.xlsx' wurde erfolgreich erstellt.")
            st.download_button(
                label="Herunterladen",
                data=output,
                file_name=f"{output_filename}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"Ein Fehler ist aufgetreten: {e}")
        st.code(traceback.format_exc())
